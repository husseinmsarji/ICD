"""Traceability matrix generator.

Emits one row per signal mapping it to its parent interface, LRUs, DAL, owning
document, and the input hash. Produced as both CSV (deterministic, diffable,
the certification primary) and XLSX (review convenience).
"""
from __future__ import annotations

import csv
import datetime
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .model import IcdModel
from .provenance import Provenance

_HEADERS = [
    "Signal", "Packet", "Interface ID", "Interface Name", "Bus Type", "DAL",
    "Source LRU", "Destination LRU", "Owning Document",
    "Signal Type", "Units", "Range Min", "Range Max", "Update Rate (Hz)",
    "Data Bits", "Xmit Bits", "Xmit Bytes", "Scale", "Offset",
    "Input SHA-256",
]


def _rows(model: IcdModel, prov: Provenance):
    for iface, pkt, sig in model.all_signals():
        yield [
            sig.name, pkt.name, iface.id, iface.name, iface.bus_type, iface.dal,
            iface.source_lru, iface.destination_lru, iface.owning_document,
            sig.signal_type, sig.units, sig.range_min, sig.range_max,
            sig.update_rate_hz,
            "" if sig.data_bits is None else sig.data_bits,
            "" if sig.xmit_bits is None else sig.xmit_bits,
            "" if sig.xmit_bytes is None else sig.xmit_bytes,
            sig.scaling, sig.offset, prov.input_hash,
        ]


def render_csv(model: IcdModel, prov: Provenance) -> str:
    buf = io.StringIO(newline="")
    # Fixed line terminator so output is byte-identical across platforms.
    writer = csv.writer(buf, lineterminator="\n")
    writer.writerow(_HEADERS)
    for row in _rows(model, prov):
        writer.writerow(row)
    return buf.getvalue()


def write_xlsx(model: IcdModel, prov: Provenance, path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Traceability"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(_HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in _rows(model, prov):
        ws.append(row)

    # Reasonable, fixed column widths (determinism: no autosize heuristics).
    widths = [22, 16, 14, 24, 12, 5, 16, 16, 20, 12, 10, 11, 11, 14, 10, 10, 11, 8, 8, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"

    # Pin workbook timestamps to a fixed epoch so identical inputs produce
    # byte-identical files (openpyxl rejects None here).
    epoch = datetime.datetime(2000, 1, 1, 0, 0, 0)
    wb.properties.created = epoch
    wb.properties.modified = epoch
    wb.properties.creator = prov.tool_name
    wb.save(path)
    from .ooxml_determinism import normalize
    normalize(path)
